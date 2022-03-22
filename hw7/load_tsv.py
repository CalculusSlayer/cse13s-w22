from openpyxl import load_workbook
import numpy as np

wb = load_workbook("myCustomers.xlsx")
sheet = wb.worksheets[0]

foods = ['apples', 'cookies', 'hotdogs', 'pho', 'donuts',
        'boba', 'cake', 'hamburgers', 'steak', 'pasta',
        'chocolate', 'brocoli', 'yogurt', 'cheese']
np.random.seed(0)

with open("customers.tsv", "w") as outfile:
    for i in range(1, sheet.max_row + 1):
        c1 = sheet.cell(row = i, column = 2).value
        c2 = sheet.cell(row = i, column = 1).value
        c3 = np.random.randint(1, 15)
        c4 = foods[np.random.randint(0, len(foods)-1)]
        print("{}\t{}\t{}\t{}".format(c1, c2, c3, c4), file=outfile)
#print(sheet.cell(row = 1, column = 1).value)
#print(type(sheet.cell(row = 1, column = 1).value))
